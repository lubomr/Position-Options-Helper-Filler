from openpyxl import load_workbook
import os
import oandapyV20
import oandapyV20.endpoints.instruments as instruments
import pandas as pd
from pandas.tseries.offsets import BDay

s= os.getcwd()              #check where script and Excecl Workbook is
for i in os.listdir():      #open the Excel workbook
    if ".xlsm" in i:
        #print(s+"\\"+i)
        WBname=i
        #print(WBname)
        wb = load_workbook(s+"\\"+i,keep_vba=True)

#print(wb.active)
ws=wb.active

client = oandapyV20.API(access_token="f7e9e4ce6d3053fd67480df1fb51e665-a52e75b6119ff67521f77ebfde0aafe0")
#Get parameters for last 20 Day candles.
params={"count":22,"granularity":"D"}

r = instruments.InstrumentsCandles(instrument="EUR_USD",params=params)
client.request(r)
#print(r.response)
#print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
for x in r.response.items():
    if not isinstance(x[1],list):
        print(x)
    #else: print(x[0])
    for y in x:
        if isinstance(y,list):
            CandlesList = y
            for i in y:
                DayCandle = i
               # print(DayCandle)
candles=[]
for candle in CandlesList:
    candles.append((candle['mid']["o"],candle['mid']["h"],candle['mid']["l"],candle['mid']["c"]))

rows=[]
for row in ws.iter_rows(min_row=3,min_col=3,max_row=24,max_col=6):
    rows.append(row)
#fill tha table

for i in range(len(candles)):
 #   print(rows[i])
 #   print(candles[i])
    for x in range(4):
        rows[i][x].value=float(candles[i][x])

#Fill the date
today = pd.datetime.today()
#print(today.date().today())
DatesList=[]
for i in range(22):
    DatesList.append(today.date().today()- BDay(i))
DatesList.reverse()
for d in range(len(DatesList)):
    ws['B'+str(3+d)]=DatesList[d]
#save the table
wb.save(filename = WBname)
